from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Q
from .models import Lote, Categoria, ConfiguracaoGlobal
from vencimento.models import Casa,Lote, HistoricoLog
from django.utils import timezone
from django.db import transaction
from django.urls import reverse
from datetime import timedelta
from django.db.models import Count
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import requests
from requests import request
from .forms import LoteForm, Lote
from vencimento.forms import TransferenciaLoteForm, SaidaProdutoForm
from django.contrib import messages
from django.core.mail import send_mail
import json


def send_notification_email(request):
    send_mail(
        'Lote Próximo do Vencimento',  # Assunto
        'Seu lote está próximo de vencer.',  # Mensagem
        'servicedesk@rioscenarium.com.br',  # E-mail de origem
        ['destinatario@example.com'],  # Lista de e-mails que receberão a mensagem
        fail_silently=False,
    )
    return HttpResponse("E-mail enviado com sucesso!")



@login_required
def lista_lotes(request):
    hoje = timezone.now().date()
    query = request.GET.get('q')
    chegada_de = request.GET.get('chegada_de')
    chegada_ate = request.GET.get('chegada_ate')
    validade_de = request.GET.get('validade_de')
    validade_ate = request.GET.get('validade_ate')
    categoria_id = request.GET.get('categoria')
    status = request.GET.get('status')
    casa_atual_id = request.session.get('casa_id', None)
    casa_atual = Casa.objects.get(id=casa_atual_id)
    configuracao = ConfiguracaoGlobal.objects.first()
    dias_para_vencimento_proximo = configuracao.dias_para_vencimento_proximo if configuracao else 30
    dias_para_urgencia = 30
    lotes = Lote.objects.filter(casa=casa_atual)  # Outras casas só veem lotes transferidos
    casa_destinada_id = request.GET.get('casa_destinada')  # ID da casa destinada
    casas = Casa.objects.all()  # Pegar todas as casas para o dropdown

    if casa_destinada_id:  # Filtrar por casa destinada se um ID foi fornecido
        lotes = lotes.filter(casa_destinada_id=casa_destinada_id)


    
    # Incluir a lógica para tratar a saída de produtos
    if request.method == 'POST':
        lote_id = request.POST.get('lote_id')
        lote = get_object_or_404(Lote, pk=lote_id)
        quantidade_a_retirar = int(request.POST.get('quantidade_a_retirar'))
        if lote.quantidade >= quantidade_a_retirar:
            lote.quantidade -= quantidade_a_retirar
            lote.save()
            messages.success(request, f"Saida de {quantidade_a_retirar} unidades registrada para o lote {lote.identificacao}.")
        else:
            messages.error(request, "Quantidade a retirar excede o estoque disponível.")
        return redirect('lista_lotes')

    if query:
        lotes = lotes.filter(Q(produto__nome__icontains=query) | Q(identificacao__icontains=query))

    if chegada_de:
        lotes = lotes.filter(data_chegada__gte=chegada_de)
    if chegada_ate:
        lotes = lotes.filter(data_chegada__lte=chegada_ate)

    if validade_de:
        lotes = lotes.filter(data_validade__gte=validade_de)
    if validade_ate:
        lotes = lotes.filter(data_validade__lte=validade_ate)
    if categoria_id:
        lotes = lotes.filter(categoria__id=categoria_id)
    
    # filtro por status
    if status == 'valido':
        lotes = lotes.filter(data_validade__gt=hoje + timedelta(days=dias_para_vencimento_proximo))
    elif status == 'proximo':
        lotes = lotes.filter(data_validade__gt=hoje + timedelta(days=dias_para_urgencia), data_validade__lte=hoje + timedelta(days=dias_para_vencimento_proximo))
    elif status == 'urgente':
        lotes = lotes.filter(data_validade__gt=hoje, data_validade__lte=hoje + timedelta(days=dias_para_urgencia))
    elif status == 'vencido':
        lotes = lotes.filter(data_validade__lte=hoje)


    categorias = Categoria.objects.all()
    categoria_data = list(lotes.values('categoria__nome')
                                .annotate(total=Count('id'))
                                .values_list('categoria__nome', 'total'))
    # Refinar as condições de status para evitar sobreposições
    status_counts = {
        'Valido': lotes.filter(data_validade__gt=hoje + timedelta(days=dias_para_vencimento_proximo)).count(),
        'Urgente': lotes.filter(data_validade__gt=hoje, data_validade__lte=hoje + timedelta(days=dias_para_urgencia)).count(),
        'Proximo do Vencimento': lotes.filter(data_validade__gt=hoje + timedelta(days=dias_para_urgencia), data_validade__lte=hoje + timedelta(days=dias_para_vencimento_proximo)).count(),
        'Vencido': lotes.filter(data_validade__lte=hoje).count(),
    }

    return render(request, 'vencimento/lista_lotes.html', {
        'lotes': lotes,
        'categorias': categorias,
        'casas': casas,  # Passar 'casas' para o contexto
        'categoria_data': json.dumps(categoria_data),
        'status_data': [(key, value) for key, value in status_counts.items()]
    })




@login_required
@require_POST
def deletar_lote_view(request):
    lote_id = request.POST.get('lote_id')
    if not lote_id:
        messages.error(request, 'ID do lote não fornecido.')
        return redirect('vencimento:lista_lotes')

    try:
        lote = Lote.objects.get(id=lote_id)
        with transaction.atomic():
            # Registro no histórico de logs antes de deletar o lote
            log = HistoricoLog.objects.create(
                usuario=request.user,
                lote=lote,
                casa=lote.casa,
                acao='DE',
                descricao=f'Deletado lote {lote.identificacao} de {lote.produto.nome} na casa {lote.casa.nome}',
            )
            lote.delete()
            messages.success(request, 'Lote deletado com sucesso.')
    except Lote.DoesNotExist:
        messages.error(request, 'Lote não encontrado.')
    except Exception as e:
        messages.error(request, f'Erro ao deletar lote: {str(e)}')

    return redirect('vencimento:lista_lotes')




@login_required
def lotes_vencidos(request):
    lotes = Lote.objects.filter(data_validade__lt=timezone.now().date())
    return render(request, 'vencimento/lista_lotes.html', {'lotes': lotes})


@login_required
def adicionar_lote(request):
    if request.method == 'POST':
        form = LoteForm(request.POST)
        if form.is_valid():
            # Passar o usuário logado para o método save
            form.save(user=request.user)
            messages.success(request, 'Lote adicionado com sucesso!')
            return redirect('vencimento:lista_lotes')
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = LoteForm()
    return render(request, 'vencimento/adicionar_lote.html', {'form': form})


@require_POST
@login_required
def registrar_saida(request):
    lote_id = request.POST.get('lote_id')
    quantidade = request.POST.get('quantidade_a_retirar')
    try:
        lote = Lote.objects.get(id=lote_id)
        if lote.quantidade >= int(quantidade):
            lote.quantidade -= int(quantidade)
            lote.save()

            # Registrar a ação no histórico de logs
            HistoricoLog.objects.create(
                usuario=request.user,
                lote=lote,
                casa=lote.casa,
                acao='RS',
                descricao=f'Registrada saída de {quantidade} unidades do lote {lote.identificacao} do produto {lote.produto.nome} na casa {lote.casa.nome}.'
            )

            messages.success(request, 'Saída registrada com sucesso.')
        else:
            messages.error(request, 'Quantidade a retirar excede o estoque disponível.')
    except Lote.DoesNotExist:
        messages.error(request, 'Lote não encontrado.')

    return redirect('vencimento:lista_lotes')


@login_required
def transferencia_lotes_view(request):
    form = TransferenciaLoteForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        result = form.save()  # Esta chamada agora retorna uma tupla se bem-sucedida
        if result:
            novo_lote, lote_original, quantidade = result
            messages.success(request, f'Lote transferido com sucesso para {novo_lote.casa.nome}.')
            
            # Registrar a transferência no histórico de logs
            HistoricoLog.objects.create(
                usuario=request.user,
                lote=lote_original,
                casa=novo_lote.casa,
                acao='TR',
                descricao=f'Transferência de {quantidade} unidades do lote {lote_original.identificacao} para a casa {novo_lote.casa.nome}.'
            )
        else:
            messages.error(request, 'Falha ao transferir lote. Verifique a quantidade disponível.')

        return redirect('vencimento:lista_lotes')

    return render(request, 'vencimento/transferencia_lotes.html', {'form': form})






@login_required
def exportar_lotes_excel(request):
    casa_id = request.session.get('casa_id')
    if casa_id:
        # Filtrar lotes pela casa em sessão
        lotes = Lote.objects.filter(casa_id=casa_id).values(
            'produto__nome', 'categoria__nome', 'identificacao',
            'quantidade', 'data_chegada', 'data_validade'
        )
    else:
        # Se não houver casa na sessão (ou se for superusuário), considera todos os lotes
        lotes = Lote.objects.all().values(
            'produto__nome', 'categoria__nome', 'identificacao',
            'quantidade', 'data_chegada', 'data_validade'
        )

    df = pd.DataFrame(lotes)
    df.columns = ['Produto', 'Categoria', 'Identificação', 'Quantidade', 'Data de Chegada', 'Data de Validade']

    # Transformar todos os dados da coluna 'Identificação' para maiúsculas
    df['Identificação'] = df['Identificação'].str.upper()

    # Garantir que as colunas de data sejam tratadas como data
    df['Data de Chegada'] = pd.to_datetime(df['Data de Chegada'])
    df['Data de Validade'] = pd.to_datetime(df['Data de Validade'])

    nome_arquivo = "lista_de_lotes.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Lotes')
        workbook = writer.book
        worksheet = writer.sheets['Lotes']

        # Formatação de cabeçalho
        header_font = Font(bold=True, color='FFFFFF')
        fill = PatternFill("solid", fgColor="4F81BD")
        for col_num, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = fill

        # Ajustar a largura das colunas
        for col in df.columns:
            max_length = max(df[col].astype(str).map(len).max(), len(col))
            worksheet.column_dimensions[get_column_letter(df.columns.get_loc(col) + 1)].width = max_length

        # Formatação de células de data
        date_format = 'dd/mm/yyyy'
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=5, max_col=6):
            for cell in row:
                cell.number_format = date_format

    return response